# Required packages:
#   pip install pandas numpy requests selenium webdriver-manager xlwings openpyxl oracledb xlsxwriter

import os
import re
import time
import json
import queue
import threading
import logging
import traceback
import calendar
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass
import tkinter as tk
from tkinter import ttk, messagebox

import numpy as np
import pandas as pd
import requests
import urllib3
import urllib3.exceptions
import xlwings as xw
import oracledb

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver import EdgeOptions
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.core.download_manager import WDMDownloadManager
from webdriver_manager.core.http import HttpClient
from webdriver_manager.microsoft import EdgeChromiumDriverManager

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

@dataclass(frozen=True)
class Config:
    DEFAULT_EDM_LINK: str = "http://edm2.sec.samsung.net/cc/link/verLink/174338140443604632/2"
    INSTANTCLIENT_DIR: str = r"C:\instantclient"
    ORACLE_HOST: str = os.getenv("ORACLE_HOST", "gmgsdd09-vip.sec.samsung.net")
    ORACLE_PORT: int = int(os.getenv("ORACLE_PORT", "2541"))
    ORACLE_SERVICE: str = os.getenv("ORACLE_SERVICE", "MEMSCM")
    ORACLE_USER: str = os.getenv("ORACLE_USER", "memscm")
    ORACLE_PW: str = os.getenv("ORACLE_PW", "mem01scm")
    HTTP_PROXY: str = "http://12.26.204.100:8080"
    HTTPS_PROXY: str = "http://12.26.204.100:8080"
    OLD_TABLE_NAME: str = "gui_eba_2yr"
    NEW_TABLE_NAME: str = "gui_eba_2yr_new"
    REQUEST_TIMEOUT: int = 30
    SETTINGS_PATH: str = "eba_settings.json"
    FAM6_MAPPING_PATH: str = "fam6_mapping.txt"

CFG = Config()

GUBUN_VALUES = [
    "ASY_EOH", "CA", "CE", "CF", "CT", "C_TOT", "EDS_EOH", "FAB_EOH", "FAB_GUIDE",
    "FAB_IN", "FAB_OUT", "FAB_P_E", "GD", "STK_EOH", "TOT_EOH", "TST_EOH", "WH_ADJ_B",
    "WH_AVAIL", "WH_EOH", "WH_EXE", "WH_GUIDE", "WH_STOCK", "WORK_DAY", "YA", "YC",
    "YE", "YF", "YL", "YT"
]

SIM_COLUMNS = [
    "FAM6", "LINE", "DESIGN_RULE", "PROD_DESC", "GUBUN",
    "PRE_M9", "PRE_M10", "PRE_M11", "PRE_M12",
    "Y0M1", "Y0M2", "Y0M3", "Y0M4", "Y0M5", "Y0M6",
    "Y0M7", "Y0M8", "Y0M9", "Y0M10", "Y0M11", "Y0M12",
    "Y1M1", "Y1M2", "Y1M3", "Y1M4", "Y1M5", "Y1M6",
    "Y1M7", "Y1M8", "Y1M9", "Y1M10", "Y1M11", "Y1M12"
]

DF_LINE = pd.DataFrame({
    "LINE": ["4", "3", "B", "P", "J", "H", "G", "E", "C", "W", "L", "M"],
    "LINE2": ["P4L", "P3L", "P2L", "P1L", "17L", "16L", "15L", "13L", "12L", "11L", "X2L", "X1L"]
})
DF_DR = pd.DataFrame({
    "DESIGN_RULE": [
        "D0A", "D1D", "D1C", "D1B-F", "D1B", "D1A", "D1Z-F", "D1Z", "D1Y-F", "D1Y", "D1X-F", "D1X",
        "D20", "D25", "V568", "V430", "V286", "V236", "V176", "V133", "V128C", "V128", "V92", "FN14-B"
    ],
    "DR2": [
        "D0A", "D1D", "D1C", "D1B", "D1B", "D1A", "D1Z", "D1Z", "D1Y", "D1Y", "D1X", "D1X",
        "D20", "D25", "V11", "V10", "V9", "V8", "V7", "V6P", "V6", "V6", "V5", "FET-B"
    ]
})
DF_DR2 = pd.DataFrame({
    "DESIGN_RULE": [
        "D0A", "D1D", "D1C", "D1B-F", "D1B", "D1A", "D1Z-F", "D1Z", "D1Y-F", "D1Y", "D1X-F", "D1X",
        "D20", "D25", "V568", "V430", "V286", "V236", "V176", "V133", "V128C", "V128", "V92", "FN14-B"
    ],
    "DR3": [
        "D0A", "D1D", "D1C", "D1B", "D1B", "D1A", "D1Z", "D1Z", "D1Y", "D1Y", "D1X", "D1X",
        "D20", "D20", "V11", "V10", "V9", "V8", "V7", "V6P", "V6", "V6", "V5", "FET-B"
    ]
})
ORDER_LINE3 = ["P4L", "P3L", "P2L", "P1L", "17L", "16L", "15L", "13L", "12L", "11L", "X2L", "X1L"]
ORDER_DR2 = ["D1D", "D1C", "D1B", "D1A", "D1Z", "D1Y", "D1X", "D20", "D25", "V11", "V10", "V9", "V8", "V7", "V6P", "V6", "V5", "FET-B"]

QUERY = """
WITH IPGO_PORTION_2627 AS (
        SELECT FAM6, LINE, YEARMONTH, YEARMONTH_IPGO,TAT_BE,
               RATIO_TO_REPORT(NUM_VALUE) OVER(PARTITION BY FAM6, LINE, YEARMONTH, TAT_BE) AS IPGO_PORTION,
               RATIO_TO_REPORT(NUM_VALUE) OVER(PARTITION BY FAM6, LINE, YEARMONTH, TAT_BE) WF_PORTION,
               CASE WHEN SUBSTR(YEARMONTH,1,4)!='2025' THEN RATIO_TO_REPORT(NUM_VALUE) OVER(PARTITION BY FAM6,  LINE, YEARMONTH, TAT_BE) ELSE 0 END WF_AVAIL_PORTION,
               CASE WHEN SUBSTR(YEARMONTH_IPGO,1,4)!='2025' THEN RATIO_TO_REPORT(NUM_VALUE) OVER(PARTITION BY FAM6,  LINE, YEARMONTH, TAT_BE) ELSE 0 END IPGO_AVAIL_PORTION
        FROM
        (
            SELECT FAM6, LINE, YEARMONTH, YEARMONTH_IPGO, COUNT(*) NUM_VALUE, max(TAT_BE) TAT_BE
            FROM
            (
                SELECT FAM6, LINE, YEARMONTH, DATEID,  TAT_BE, TO_CHAR(TAT_BE+DATEID,'yyyymm') YEARMONTH_IPGO
                FROM
                (
                    SELECT  A.FAM6_ADJ FAM6, YEARMONTH,
                            LINE,
                            CASE
                                WHEN LINE = '4' THEN 'PFB4'
                                WHEN LINE = '3' THEN 'PFB3'
                                WHEN LINE = 'B' THEN 'PFBB'
                                WHEN LINE = 'P' THEN 'PFBP'
                                WHEN LINE = 'J' THEN 'KFBJ'
                                WHEN LINE = 'H' THEN 'KFBH'
                                WHEN LINE = 'G' THEN 'KFBG'
                                WHEN LINE = 'E' THEN 'KFBE'
                                WHEN LINE = 'C' THEN 'KFBC'
                                WHEN LINE = 'W' THEN 'KFBW'
                                WHEN LINE = 'L' THEN 'XFB2'
                                WHEN LINE = 'M' THEN 'XFB1'
                            ELSE '-'
                            END AS LINE_CODE
                          , DECODE(GUBUN,'C_TOT',DECODE(ORDER_NUM,-3, PRE_M9, -2,PRE_M10,-1,PRE_M11,0,PRE_M12, 1,Y0M1, 2,Y0M2, 3, Y0M3, 4,Y0M4, 5,Y0M5, 6,Y0M6, 7,Y0M7, 8,Y0M8, 9,Y0M9, 10,Y0M10, 11,Y0M11, 12,Y0M12, 13, Y1M1, 14,Y1M2, 15, Y1M3, 16,Y1M4, 17,Y1M5, 18,Y1M6, 19,Y1M7, 20,Y1M8, 21,Y1M9, 22,Y1M10, 23,Y1M11, 24,Y1M12,0),0) TAT_BE
                    FROM gui_eba_2yr_test A
                    , (
                            SELECT YEARMONTH, ROW_NUMBER() OVER( ORDER BY YEARMONTH)-4 ORDER_NUM
                            FROM
                            (
                                SELECT DISTINCT FNCMONTH YEARMONTH FROM MST_FNC_TIME
                                WHERE FNCMONTH >= '202509'
                                AND FNCMONTH <= '202712'
                            )
                    ) B
                    WHERE 1=1
                    AND PLANID = :plan_id
                    AND GUBUN IN ('C_TOT')
                ) P, MST_FNC_TIME Q
                WHERE P.YEARMONTH=Q.FNCMONTH
                AND FNCMONTH>='202509'
                AND  FNCMONTH<='202712'
            )
            GROUP BY FAM6, LINE, YEARMONTH, YEARMONTH_IPGO
            ORDER BY FAM6, LINE, YEARMONTH
        )
)
SELECT X.PLANID, X.FAM1, X.FAM5, X.LINE, X.DESIGN_RULE, X.FAM6, X.VERSION, X.DR, SUBSTR(X.YEARMONTH,1,4) YEAR, X.YEARMONTH, X.NETDIE, Z.TG, X.EQ, WF_TTL, NVL(PKG_OUT_EQ/100000,0) PKG입고_억EQ, NVL(순입고_MEQ_FINAL/100,0) 순입고_억EQ
, WF_P_E
, NVL(ROUND((WF_TTL-WF_P_E)*X.NETDIE*Z.TG*X.EQ/100000000,7),0) 순생산_억EQ, TAT_BE
FROM (
    SELECT PLANID, FAM1, FAM5, LINE, DESIGN_RULE, FAM6, VERSION, DR, YEARMONTH, NETDIE, EQ, SUM(WF_OUT) WF_TTL, SUM(PKG_OUT) PKG_OUT_CHIP, SUM(PKG_OUT*EQ) PKG_OUT_EQ, SUM(WF_P_E) WF_P_E
    FROM (
            SELECT  PLANID, FAM1, FAM5, P.LINE, DESIGN_RULE, P.FAM6_ADJ FAM6, VERSION, DR, YEARMONTH, NETDIE, EQ
                  , DECODE(GUBUN,'FAB_OUT',DECODE(ORDER_NUM,-3, PRE_M9, -2,PRE_M10,-1,PRE_M11,0,PRE_M12, 1,Y0M1, 2,Y0M2, 3, Y0M3, 4,Y0M4, 5,Y0M5, 6,Y0M6, 7,Y0M7, 8,Y0M8, 9,Y0M9, 10,Y0M10, 11,Y0M11, 12,Y0M12, 13, Y1M1, 14,Y1M2, 15, Y1M3, 16,Y1M4, 17,Y1M5, 18,Y1M6, 19,Y1M7, 20,Y1M8, 21,Y1M9, 22,Y1M10, 23,Y1M11, 24,Y1M12,0),0) WF_OUT
                  , DECODE(GUBUN,'WH_GUIDE',DECODE(ORDER_NUM,-3, PRE_M9, -2,PRE_M10,-1,PRE_M11,0,PRE_M12, 1,Y0M1, 2,Y0M2, 3, Y0M3, 4,Y0M4, 5,Y0M5, 6,Y0M6, 7,Y0M7, 8,Y0M8, 9,Y0M9, 10,Y0M10, 11,Y0M11, 12,Y0M12, 13, Y1M1, 14,Y1M2, 15, Y1M3, 16,Y1M4, 17,Y1M5, 18,Y1M6, 19,Y1M7, 20,Y1M8, 21,Y1M9, 22,Y1M10, 23,Y1M11, 24,Y1M12,0),0) PKG_OUT
                  , DECODE(GUBUN,'FAB_P_E',DECODE(ORDER_NUM,-3, PRE_M9, -2,PRE_M10,-1,PRE_M11,0,PRE_M12, 1,Y0M1, 2,Y0M2, 3, Y0M3, 4,Y0M4, 5,Y0M5, 6,Y0M6, 7,Y0M7, 8,Y0M8, 9,Y0M9, 10,Y0M10, 11,Y0M11, 12,Y0M12, 13, Y1M1, 14,Y1M2, 15, Y1M3, 16,Y1M4, 17,Y1M5, 18,Y1M6, 19,Y1M7, 20,Y1M8, 21,Y1M9, 22,Y1M10, 23,Y1M11, 24,Y1M12,0),0) WF_P_E
            FROM gui_eba_2yr_test P
            , (
                    SELECT YEARMONTH, ROW_NUMBER() OVER( ORDER BY YEARMONTH)-4 ORDER_NUM
                    FROM
                    (
                        SELECT DISTINCT FNCMONTH YEARMONTH FROM MST_FNC_TIME
                        WHERE FNCMONTH >= '202509'
                        AND FNCMONTH <= '202712'
                    )
            ) Q
            ,(
                  SELECT FAM6, MAX(VERSION) VERSION, MAX(DR) DR, MAX(FAM1) FAM1, MAX(EQ) EQ, MAX(FAM5) FAM5, MAX(NETDIE) NETDIE
                  FROM  MEMBP_SDB.EXP_MST_BP_MASTER
                  WHERE WORK_FLAG ='Y'
                  GROUP BY FAM6
            ) R
            WHERE 1=1
            AND PLANID = :plan_id
            AND GUBUN IN ('FAB_OUT','FAB_P_E','WH_GUIDE')
            AND P.FAM6_ADJ = R.FAM6 (+)
    )
   GROUP BY PLANID, FAM1, FAM5, LINE, DESIGN_RULE, FAM6, VERSION, DR, YEARMONTH, NETDIE, EQ
) X
,(
        SELECT FAM1, FAM5, LINE, DESIGN_RULE, FAM6, VERSION, YEARMONTH_IPGO, 순입고_MEQ_ORG, TAT_BE
            , CASE WHEN FAM6 LIKE '%HBM%' THEN 순입고_MEQ_ORG ELSE 순입고_MEQ_ORG END 순입고_MEQ_FINAL
        FROM (
                SELECT  A.FAM1, A.FAM5, A.LINE, A.DESIGN_RULE, A.FAM6, A.VERSION, YEARMONTH_IPGO, EQ, SUM(WF_TTL) WF_OUT, SUM(PKG_OUT) PKG입고_CHIP
                       , ROUND(SUM(NVL(ROUND(WF_OUT*IPGO_AVAIL_PORTION*NETDIE*TOTAL_GROSS*EQ,8)/1000000,0)),5) 순입고_MEQ_ORG
                       , ROUND(SUM(PKG_OUT*EQ)/1000) PKG입고_MEQ, max(TAT_BE) TAT_BE
                FROM (
                    SELECT FAM1, FAM5, LINE, DESIGN_RULE, FAM6, VERSION, DR, YEARMONTH, NETDIE, EQ, SUM(WF_OUT) WF_TTL, SUM(WF_OUT-WF_ES_SAMPLE) WF_OUT, SUM(PKG_OUT) PKG_OUT
                    FROM (
                        SELECT  FAM1, FAM5, LINE, DESIGN_RULE, FAM6_ADJ FAM6, VERSION, DR, YEARMONTH, NETDIE, EQ
                              , DECODE(GUBUN,'FAB_OUT',DECODE(ORDER_NUM,-3, PRE_M9, -2,PRE_M10,-1,PRE_M11,0,PRE_M12, 1,Y0M1, 2,Y0M2, 3, Y0M3, 4,Y0M4, 5,Y0M5, 6,Y0M6, 7,Y0M7, 8,Y0M8, 9,Y0M9, 10,Y0M10, 11,Y0M11, 12,Y0M12, 13, Y1M1, 14,Y1M2, 15, Y1M3, 16,Y1M4, 17,Y1M5, 18,Y1M6, 19,Y1M7, 20,Y1M8, 21,Y1M9, 22,Y1M10, 23,Y1M11, 24,Y1M12,0),0) WF_OUT
                              , DECODE(GUBUN,'WH_GUIDE',DECODE(ORDER_NUM,-3, PRE_M9, -2,PRE_M10,-1,PRE_M11,0,PRE_M12, 1,Y0M1, 2,Y0M2, 3, Y0M3, 4,Y0M4, 5,Y0M5, 6,Y0M6, 7,Y0M7, 8,Y0M8, 9,Y0M9, 10,Y0M10, 11,Y0M11, 12,Y0M12, 13, Y1M1, 14,Y1M2, 15, Y1M3, 16,Y1M4, 17,Y1M5, 18,Y1M6, 19,Y1M7, 20,Y1M8, 21,Y1M9, 22,Y1M10, 23,Y1M11, 24,Y1M12,0),0) PKG_OUT
                              , DECODE(GUBUN,'FAB_P_E',DECODE(ORDER_NUM,-3, PRE_M9, -2,PRE_M10,-1,PRE_M11,0,PRE_M12, 1,Y0M1, 2,Y0M2, 3, Y0M3, 4,Y0M4, 5,Y0M5, 6,Y0M6, 7,Y0M7, 8,Y0M8, 9,Y0M9, 10,Y0M10, 11,Y0M11, 12,Y0M12, 13, Y1M1, 14,Y1M2, 15, Y1M3, 16,Y1M4, 17,Y1M5, 18,Y1M6, 19,Y1M7, 20,Y1M8, 21,Y1M9, 22,Y1M10, 23,Y1M11, 24,Y1M12,0),0) WF_ES_SAMPLE
                        FROM gui_eba_2yr_test P
                        , (
                            SELECT YEARMONTH, ROW_NUMBER() OVER( ORDER BY YEARMONTH)-4 ORDER_NUM
                            FROM (
                                SELECT DISTINCT FNCMONTH YEARMONTH FROM MST_FNC_TIME
                                WHERE FNCMONTH >= '202509'
                                AND FNCMONTH <= '202712'
                            )
                        ) Q
                        ,(
                            SELECT FAM6, MAX(VERSION) VERSION, MAX(DR) DR, MAX(FAM1) FAM1, MAX(EQ) EQ, MAX(FAM5) FAM5, MAX(NETDIE) NETDIE
                            FROM  MEMBP_SDB.EXP_MST_BP_MASTER
                            WHERE WORK_FLAG ='Y'
                            GROUP BY FAM6
                        ) R
                        WHERE 1=1
                        AND PLANID = :plan_id
                        AND GUBUN IN ('FAB_OUT','FAB_P_E','WH_GUIDE')
                        AND P.FAM6_ADJ = R.FAM6 (+)
                    )
                    GROUP BY FAM1, FAM5, LINE, DESIGN_RULE, FAM6, VERSION, DR, YEARMONTH, NETDIE, EQ
                ) A
                ,(
                    SELECT  FAM6, LINE, YEARMONTH, YEARMONTH_IPGO, IPGO_AVAIL_PORTION, TAT_BE
                    FROM IPGO_PORTION_2627
                ) B
                ,(
                    select  FAM6_ADJ FAM6, line, YEARMONTH
                           , DECODE(GUBUN,'YC',DECODE(ORDER_NUM,-3, PRE_M9, -2,PRE_M10,-1,PRE_M11,0,PRE_M12, 1,Y0M1, 2,Y0M2, 3, Y0M3, 4,Y0M4, 5,Y0M5, 6,Y0M6, 7,Y0M7, 8,Y0M8, 9,Y0M9, 10,Y0M10, 11,Y0M11, 12,Y0M12, 13, Y1M1, 14,Y1M2, 15, Y1M3, 16,Y1M4, 17,Y1M5, 18,Y1M6, 19,Y1M7, 20,Y1M8, 21,Y1M9, 22,Y1M10, 23,Y1M11, 24,Y1M12,0),0) TOTAL_GROSS
                    from gui_eba_2yr_test a,
                    (
                        SELECT YEARMONTH, ROW_NUMBER() OVER( ORDER BY YEARMONTH)-4 ORDER_NUM
                        FROM
                        (
                            SELECT DISTINCT FNCMONTH YEARMONTH FROM MST_FNC_TIME
                            WHERE FNCMONTH >= '202509'
                            AND FNCMONTH <= '202712'
                        )
                    ) B
                    WHERE 1=1
                    and PLANID = :plan_id
                    and GUBUN in ('YC')
                ) C
                WHERE A.YEARMONTH = B.YEARMONTH (+)
                AND A.FAM6 = B.FAM6 (+)
                AND A.LINE = B.LINE (+)
                AND A.FAM6 = C.FAM6
                AND A.LINE = C.LINE
                AND A.YEARMONTH = C.YEARMONTH
                AND A.FAM6 NOT LIKE 'ER%'
                GROUP BY  A.FAM1, A.FAM5, A.LINE, A.DESIGN_RULE, A.FAM6, A.VERSION, YEARMONTH_IPGO, EQ
                HAVING YEARMONTH_IPGO NOT LIKE '2025%'
        )
) Y
,(
    SELECT  FAM6_ADJ FAM6, LINE, YEARMONTH
           , DECODE(GUBUN,'YC',DECODE(ORDER_NUM,-3, PRE_M9, -2,PRE_M10,-1,PRE_M11,0,PRE_M12, 1,Y0M1, 2,Y0M2, 3, Y0M3, 4,Y0M4, 5,Y0M5, 6,Y0M6, 7,Y0M7, 8,Y0M8, 9,Y0M9, 10,Y0M10, 11,Y0M11, 12,Y0M12, 13, Y1M1, 14,Y1M2, 15, Y1M3, 16,Y1M4, 17,Y1M5, 18,Y1M6, 19,Y1M7, 20,Y1M8, 21,Y1M9, 22,Y1M10, 23,Y1M11, 24,Y1M12,0),0) TG
    FROM gui_eba_2yr_test A,
    (
        SELECT YEARMONTH, ROW_NUMBER() OVER( ORDER BY YEARMONTH)-4 ORDER_NUM
        FROM
        (
            SELECT DISTINCT FNCMONTH YEARMONTH FROM MST_FNC_TIME
            WHERE FNCMONTH >= '202509'
            AND FNCMONTH <= '202712'
        )
    ) B
    WHERE 1=1
    AND PLANID = :plan_id
    AND GUBUN IN ('YC')
    AND FAM6_ADJ NOT LIKE 'ER_%'
) Z
WHERE X.FAM6 = Y.FAM6 (+)
AND X.LINE = Y.LINE (+)
AND X.YEARMONTH = Y.YEARMONTH_IPGO (+)
AND X.FAM6 = Z.FAM6 (+)
AND X.LINE = Z. LINE (+)
AND X.YEARMONTH = Z.YEARMONTH (+)
AND X.YEARMONTH NOT LIKE '2025%'
"""

QUERY2 = """
SELECT FAM6, MAX(VERSION) VERSION, MAX(DR) DR, MAX(FAM1) FAM1, max(USER_FAM1) USER_FAM1, max(USER_FAM2) USER_FAM2 ,MAX(EQ) EQ, MAX(FAM5) FAM5, MAX(NETDIE) NETDIE
FROM  MEMBP_SDB.EXP_MST_BP_MASTER
WHERE WORK_FLAG ='Y'
GROUP BY FAM6
"""

QUERY = QUERY.replace("gui_eba_2yr_test", CFG.NEW_TABLE_NAME)

O9_BASE_QUERY = f"""
WITH BASE AS (
    SELECT A.PLANID AS VERSION_NAME, A.FAM6_ADJ, A.LINE, NVL(B.SITEID, A.LINE) AS SITEID, A.GUBUN,
           A.Y0M1, A.Y0M2, A.Y0M3, A.Y0M4, A.Y0M5, A.Y0M6, A.Y0M7, A.Y0M8, A.Y0M9, A.Y0M10, A.Y0M11, A.Y0M12,
           A.Y1M1, A.Y1M2, A.Y1M3, A.Y1M4, A.Y1M5, A.Y1M6, A.Y1M7, A.Y1M8, A.Y1M9, A.Y1M10, A.Y1M11, A.Y1M12
    FROM {CFG.NEW_TABLE_NAME} A
    LEFT JOIN MTA_FABLINE B ON A.LINE = B.LINE
    WHERE A.PLANID = :plan_id
)
SELECT VERSION_NAME, FAM6_ADJ AS FAM6, SITEID, MEASURE,
       Y0M1, Y0M2, Y0M3, Y0M4, Y0M5, Y0M6, Y0M7, Y0M8, Y0M9, Y0M10, Y0M11, Y0M12,
       Y1M1, Y1M2, Y1M3, Y1M4, Y1M5, Y1M6, Y1M7, Y1M8, Y1M9, Y1M10, Y1M11, Y1M12
FROM (
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'FABOutPlanOverride' MEASURE,
           Y0M1, Y0M2, Y0M3, Y0M4, Y0M5, Y0M6, Y0M7, Y0M8, Y0M9, Y0M10, Y0M11, Y0M12,
           Y1M1, Y1M2, Y1M3, Y1M4, Y1M5, Y1M6, Y1M7, Y1M8, Y1M9, Y1M10, Y1M11, Y1M12
    FROM BASE WHERE GUBUN = 'FAB_OUT'
    UNION ALL
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'PETransferOverride' MEASURE,
           Y0M1, Y0M2, Y0M3, Y0M4, Y0M5, Y0M6, Y0M7, Y0M8, Y0M9, Y0M10, Y0M11, Y0M12,
           Y1M1, Y1M2, Y1M3, Y1M4, Y1M5, Y1M6, Y1M7, Y1M8, Y1M9, Y1M10, Y1M11, Y1M12
    FROM BASE WHERE GUBUN = 'FAB_P_E' AND FAM6_ADJ NOT LIKE 'ER%'
    UNION ALL
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'WHInPlanOverride' MEASURE,
           Y0M1, Y0M2, Y0M3, Y0M4, Y0M5, Y0M6, Y0M7, Y0M8, Y0M9, Y0M10, Y0M11, Y0M12,
           Y1M1, Y1M2, Y1M3, Y1M4, Y1M5, Y1M6, Y1M7, Y1M8, Y1M9, Y1M10, Y1M11, Y1M12
    FROM BASE WHERE GUBUN = 'WH_GUIDE' AND FAM6_ADJ NOT LIKE 'ER%'
    UNION ALL
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'FABTATOverrideTwin' MEASURE,
           Y0M1, Y0M2, Y0M3, Y0M4, Y0M5, Y0M6, Y0M7, Y0M8, Y0M9, Y0M10, Y0M11, Y0M12,
           Y1M1, Y1M2, Y1M3, Y1M4, Y1M5, Y1M6, Y1M7, Y1M8, Y1M9, Y1M10, Y1M11, Y1M12
    FROM BASE WHERE GUBUN = 'CF'
    UNION ALL
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'EDSTATOverrideTwin' MEASURE,
           Y0M1, Y0M2, Y0M3, Y0M4, Y0M5, Y0M6, Y0M7, Y0M8, Y0M9, Y0M10, Y0M11, Y0M12,
           Y1M1, Y1M2, Y1M3, Y1M4, Y1M5, Y1M6, Y1M7, Y1M8, Y1M9, Y1M10, Y1M11, Y1M12
    FROM BASE WHERE GUBUN = 'CE' AND FAM6_ADJ NOT LIKE 'ER%'
    UNION ALL
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'BETATOverrideTwin' MEASURE,
           SUM(Y0M1), SUM(Y0M2), SUM(Y0M3), SUM(Y0M4), SUM(Y0M5), SUM(Y0M6), SUM(Y0M7), SUM(Y0M8), SUM(Y0M9), SUM(Y0M10), SUM(Y0M11), SUM(Y0M12),
           SUM(Y1M1), SUM(Y1M2), SUM(Y1M3), SUM(Y1M4), SUM(Y1M5), SUM(Y1M6), SUM(Y1M7), SUM(Y1M8), SUM(Y1M9), SUM(Y1M10), SUM(Y1M11), SUM(Y1M12)
    FROM BASE WHERE GUBUN IN ('CA', 'CT') AND FAM6_ADJ NOT LIKE 'ER%'
    GROUP BY VERSION_NAME, FAM6_ADJ, LINE, SITEID
    UNION ALL
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'FABYieldOverrideTwin' MEASURE,
           Y0M1, Y0M2, Y0M3, Y0M4, Y0M5, Y0M6, Y0M7, Y0M8, Y0M9, Y0M10, Y0M11, Y0M12,
           Y1M1, Y1M2, Y1M3, Y1M4, Y1M5, Y1M6, Y1M7, Y1M8, Y1M9, Y1M10, Y1M11, Y1M12
    FROM BASE WHERE GUBUN = 'YF'
    UNION ALL
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'EDSYieldOverrideTwin' MEASURE,
           Y0M1, Y0M2, Y0M3, Y0M4, Y0M5, Y0M6, Y0M7, Y0M8, Y0M9, Y0M10, Y0M11, Y0M12,
           Y1M1, Y1M2, Y1M3, Y1M4, Y1M5, Y1M6, Y1M7, Y1M8, Y1M9, Y1M10, Y1M11, Y1M12
    FROM BASE WHERE GUBUN = 'YE' AND FAM6_ADJ NOT LIKE 'ER%'
    UNION ALL
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'ASYYieldOverrideTwin' MEASURE,
           Y0M1, Y0M2, Y0M3, Y0M4, Y0M5, Y0M6, Y0M7, Y0M8, Y0M9, Y0M10, Y0M11, Y0M12,
           Y1M1, Y1M2, Y1M3, Y1M4, Y1M5, Y1M6, Y1M7, Y1M8, Y1M9, Y1M10, Y1M11, Y1M12
    FROM BASE WHERE GUBUN = 'YA' AND FAM6_ADJ NOT LIKE 'ER%'
    UNION ALL
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'TSTYieldOverrideTwin' MEASURE,
           Y0M1, Y0M2, Y0M3, Y0M4, Y0M5, Y0M6, Y0M7, Y0M8, Y0M9, Y0M10, Y0M11, Y0M12,
           Y1M1, Y1M2, Y1M3, Y1M4, Y1M5, Y1M6, Y1M7, Y1M8, Y1M9, Y1M10, Y1M11, Y1M12
    FROM BASE WHERE GUBUN = 'YT' AND FAM6_ADJ NOT LIKE 'ER%'
    UNION ALL
    SELECT VERSION_NAME, FAM6_ADJ, LINE, SITEID, 'MODYieldOverrideTwin' MEASURE,
           1,1,1,1,1,1,1,1,1,1,1,1, 1,1,1,1,1,1,1,1,1,1,1,1
    FROM BASE WHERE GUBUN = 'YT' AND FAM6_ADJ NOT LIKE 'ER%'
)
ORDER BY FAM6, SITEID, MEASURE DESC
"""

def build_o9_months(planid: str) -> list[str]:
    m = re.search(r"(20\d{2})(0[1-9]|1[0-2])", planid)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
    else:
        year, month = 2026, 1
    months = []
    for _ in range(24):
        months.append(f"{year}{month:02d}")
        month += 1
        if month == 13:
            month = 1
            year += 1
    return months

def sanitize_filename(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', '_', value.strip())

def load_settings() -> dict:
    path = Path(CFG.SETTINGS_PATH)
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}

def save_settings(data: dict):
    Path(CFG.SETTINGS_PATH).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def load_fam6_mapping_file() -> pd.DataFrame:
    path = Path(CFG.FAM6_MAPPING_PATH)
    if path.exists():
        raw_text = None
        last_err = None
        for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr", "latin1"]:
            try:
                raw_text = path.read_text(encoding=enc)
                break
            except Exception as e:
                last_err = e
        if raw_text is None:
            raise RuntimeError(f"FAM6 매핑 TXT를 읽지 못했습니다: {path} ({last_err})")
        rows = []
        for line in raw_text.splitlines():
            line = line.strip()
            if not line:
                continue
            if "\t" in line:
                parts = [p.strip() for p in line.split("\t")]
            elif "|" in line:
                parts = [p.strip() for p in line.split("|")]
            elif "," in line:
                parts = [p.strip() for p in line.split(",")]
            else:
                continue
            if len(parts) < 2:
                continue
            if parts[0].upper() == "FAM6" or set(parts[0]) <= {"-"}:
                continue
            rows.append((parts[0], parts[1]))
        if rows:
            return pd.DataFrame(rows, columns=["FAM6", "FAM6_ADJ"]).dropna(how="all")
    return pd.DataFrame(columns=["FAM6", "FAM6_ADJ"])

def save_fam6_mapping_file(df: pd.DataFrame):
    df = df[["FAM6", "FAM6_ADJ"]].fillna("")
    lines = ["FAM6\tFAM6_ADJ"] + [f"{row.FAM6}\t{row.FAM6_ADJ}" for row in df.itertuples(index=False)]
    Path(CFG.FAM6_MAPPING_PATH).write_text("\n".join(lines), encoding="utf-8-sig")

def month_to_quarter(month: int):
    if 1 <= month <= 3:
        return 'Q1'
    if 4 <= month <= 6:
        return 'Q2'
    if 7 <= month <= 9:
        return 'Q3'
    if 10 <= month <= 12:
        return 'Q4'
    return None

def elapsed_text(start_time: datetime) -> str:
    end_time = datetime.now()
    elapsed = end_time - start_time
    days = elapsed.days
    hours, remainder = divmod(elapsed.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"시작: {start_time:%Y-%m-%d %H:%M:%S}\n종료: {end_time:%Y-%m-%d %H:%M:%S}\n총 소요: {days}일 {hours}시간 {minutes}분 {seconds}초"

def make_df_pv(df, col_value='WF_TTL', order_line=ORDER_LINE3, colsum_true=True):
    df_pv = pd.pivot_table(df, values=col_value, index=['FAM1', 'LINE2', 'DR2'], columns='YEARMONTH', aggfunc='sum', fill_value=0, margins=True, observed=True)
    df_pv = df_pv[df_pv['All'] != 0]
    df_pv = df_pv.drop(index='All', errors='ignore')
    df_pv = df_pv.drop(columns='All', errors='ignore')
    subtotals = df_pv.groupby(level=['FAM1', 'LINE2']).sum()
    subtotals.index = pd.MultiIndex.from_tuples([(idx[0], idx[1], '합계') for idx in subtotals.index])
    df_pv = pd.concat([subtotals, df_pv])
    df_pv = df_pv.sort_index(level=[0, 1, 2], ascending=[True, True, False])
    df_pv = df_pv.reset_index()
    df_pv.columns = ['FAM1', 'LINE', 'DR'] + df_pv.columns.to_list()[3:]
    df_pv['LINE'] = pd.Categorical(df_pv['LINE'], categories=['합계'] + order_line, ordered=True)
    df_pv['DR'] = pd.Categorical(df_pv['DR'], categories=['합계'] + ORDER_DR2, ordered=True)
    df_pv = df_pv.sort_values(by=['FAM1', 'LINE', 'DR'], ascending=[True, True, True])
    if colsum_true:
        key_cols = df_pv.columns.to_list()[:3]
        columns_2026 = [col for col in df_pv.columns if str(col).startswith('2026')]
        columns_2027 = [col for col in df_pv.columns if str(col).startswith('2027')]
        df_pv['2026합'] = df_pv[columns_2026].sum(axis=1)
        df_pv['2027합'] = df_pv[columns_2027].sum(axis=1)
        df_pv = pd.concat([df_pv[key_cols], df_pv[columns_2026], df_pv[['2026합']], df_pv[columns_2027], df_pv[['2027합']]], axis=1)
    return df_pv

class QueueLogHandler(logging.Handler):
    def __init__(self, q):
        super().__init__()
        self.q = q
    def emit(self, record):
        self.q.put(self.format(record))

class CustomHttpClient(HttpClient):
    def get(self, url, params=None, **kwargs) -> requests.Response:
        return NetworkClient().get(url, params=params, **kwargs)

class NetworkClient:
    def __init__(self, logger=None):
        self.logger = logger or logging.getLogger("eba_tool")
        self.session = requests.Session()
        self.session.proxies.update({"http": CFG.HTTP_PROXY, "https": CFG.HTTPS_PROXY})
        self.session.verify = False

    def get(self, url, params=None, **kwargs):
        timeout = kwargs.pop("timeout", CFG.REQUEST_TIMEOUT)
        try:
            resp = self.session.get(url, params=params, timeout=timeout, **kwargs)
            resp.raise_for_status()
            return resp
        except Exception:
            self.logger.exception("GET 요청 실패: %s", url)
            raise

    def post(self, url, data=None, json=None, **kwargs):
        timeout = kwargs.pop("timeout", CFG.REQUEST_TIMEOUT)
        try:
            resp = self.session.post(url, data=data, json=json, timeout=timeout, **kwargs)
            resp.raise_for_status()
            return resp
        except Exception:
            self.logger.exception("POST 요청 실패: %s", url)
            raise

class OracleService:
    def __init__(self, logger):
        self.logger = logger
        self.conn = None
    def connect(self):
        try:
            oracledb.init_oracle_client(lib_dir=CFG.INSTANTCLIENT_DIR)
        except Exception:
            pass
        dsn = f"{CFG.ORACLE_HOST}:{CFG.ORACLE_PORT}/{CFG.ORACLE_SERVICE}"
        self.conn = oracledb.connect(user=CFG.ORACLE_USER, password=CFG.ORACLE_PW, dsn=dsn)
        self.logger.info('Oracle 연결 완료')
    def close(self):
        if self.conn:
            self.conn.close()
    def delete_table_if_exists(self, table_name):
        cur = self.conn.cursor()
        try:
            cur.execute(f"DROP TABLE {table_name}")
            self.conn.commit()
            self.logger.info('기존 테이블 삭제 완료: %s', table_name)
        except oracledb.DatabaseError as e:
            if 'ORA-00942' in str(e):
                self.logger.info('삭제 대상 테이블 없음: %s', table_name)
            else:
                raise
        finally:
            cur.close()
    def copy_table_structure(self, old_table_name, new_table_name):
        cur = self.conn.cursor()
        try:
            cur.execute(f"CREATE TABLE {new_table_name} AS SELECT * FROM {old_table_name} WHERE 1=0")
            self.conn.commit()
            self.logger.info('테이블 구조 복사 완료: %s -> %s', old_table_name, new_table_name)
        finally:
            cur.close()
    def insert_dataframe_into_table(self, df, table_name, batch_size=1000):
        cur = self.conn.cursor()
        work_df = df.copy()
        type_cur = self.conn.cursor()
        type_cur.execute(
            "SELECT COLUMN_NAME, DATA_TYPE FROM USER_TAB_COLUMNS WHERE TABLE_NAME = :tbl ORDER BY COLUMN_ID",
            {"tbl": table_name.upper()},
        )
        type_map = {name.upper(): dtype.upper() for name, dtype in type_cur.fetchall()}
        type_cur.close()
        for col in work_df.columns:
            dtype = type_map.get(col.upper(), "")
            if dtype in {"VARCHAR2", "VARCHAR", "CHAR", "NCHAR", "NVARCHAR2", "CLOB"}:
                work_df[col] = work_df[col].apply(lambda v: None if pd.isna(v) else str(v))
            elif dtype in {"NUMBER", "FLOAT", "BINARY_FLOAT", "BINARY_DOUBLE"}:
                work_df[col] = pd.to_numeric(work_df[col], errors="coerce")
            elif dtype == "DATE" or dtype.startswith("TIMESTAMP"):
                work_df[col] = pd.to_datetime(work_df[col], errors="coerce")
                work_df[col] = work_df[col].apply(
                    lambda v: None if pd.isna(v) else (v.to_pydatetime() if isinstance(v, pd.Timestamp) else v)
                )
        work_df = work_df.replace({np.nan: None}).where(pd.notnull(work_df), None)
        cols = list(work_df.columns)
        sql = f"INSERT INTO {table_name} ({', '.join(cols)}) VALUES ({', '.join(f':{i+1}' for i in range(len(cols)))})"
        rows = [tuple(r) for r in work_df.itertuples(index=False, name=None)]
        try:
            for start in range(0, len(rows), batch_size):
                batch = rows[start:start+batch_size]
                cur.executemany(sql, batch)
                self.conn.commit()
                self.logger.info('DB 적재 진행: %s / %s', min(start + batch_size, len(rows)), len(rows))
        except oracledb.NotSupportedError:
            self.logger.warning("executemany 타입 바인딩 실패, 행 단위 적재로 재시도합니다.")
            self.conn.rollback()
            for idx, row in enumerate(rows, 1):
                cur.execute(sql, row)
                if idx % batch_size == 0 or idx == len(rows):
                    self.conn.commit()
                    self.logger.info('DB 적재 진행(행단위): %s / %s', idx, len(rows))
        cur.close()
    def read_sql(self, sql, params=None):
        return pd.read_sql(sql, self.conn, params=params)

class EDMClient:
    def __init__(self, logger):
        self.logger = logger
    def build_driver(self):
        dm = WDMDownloadManager(CustomHttpClient())
        options = EdgeOptions()
        options.add_argument('--start-maximized')
        service = EdgeService(EdgeChromiumDriverManager(download_manager=dm, url='https://msedgedriver.microsoft.com', latest_release_url='https://msedgedriver.microsoft.com/LATEST_RELEASE').install())
        return webdriver.Edge(service=service, options=options)
    def _snapshot_books(self):
        names = set()
        try:
            for app in xw.apps:
                for book in app.books:
                    names.add((app.pid, book.name))
        except Exception:
            pass
        return names
    def _is_target_workbook(self, book):
        names = {s.name for s in book.sheets}
        return ('Simulation' in names) or any('Simulation' in s for s in names)
    def _wait_new_workbook(self, before, timeout=120):
        start = time.time()
        while time.time() - start < timeout:
            for app in xw.apps:
                for book in app.books:
                    key = (app.pid, book.name)
                    if key not in before and self._is_target_workbook(book):
                        return book
            for app in xw.apps:
                for book in app.books:
                    if self._is_target_workbook(book):
                        return book
            time.sleep(2)
        raise RuntimeError('EDM에서 열린 Excel 워크북을 찾지 못했습니다.')
    def open_edm_and_attach_workbook(self, edm_link):
        before = self._snapshot_books()
        driver = self.build_driver()
        try:
            driver.get(edm_link)
            button_selectors = [
                (By.CSS_SELECTOR, "div.btns span.r a:nth-child(2)"),
                (By.XPATH, "//a[contains(.,'보기') or contains(.,'열기') or contains(.,'Open') or contains(.,'View')]"),
                (By.XPATH, "//button[contains(.,'보기') or contains(.,'열기') or contains(.,'Open') or contains(.,'View')]"),
            ]
            clicked = False
            for by, selector in button_selectors:
                try:
                    element = WebDriverWait(driver, timeout=8).until(EC.element_to_be_clickable((by, selector)))
                    self.logger.info('EDM 버튼 탐지: %s', element.text)
                    element.click()
                    clicked = True
                    break
                except Exception:
                    continue
            if not clicked:
                raise RuntimeError("EDM 화면에서 보기/열기 버튼을 찾지 못했습니다.")
            time.sleep(5)
        finally:
            driver.quit()
        wb = self._wait_new_workbook(before)
        self.logger.info('Excel 워크북 연결 완료: %s', wb.name)
        return wb
    def read_simulation_and_fam6(self, wb):
        sim_sheet = next((sh for sh in wb.sheets if 'Simulation' in sh.name), None)
        if sim_sheet is None:
            raise RuntimeError('Simulation 시트를 찾지 못했습니다.')
        fam6_sheet = next((sh for sh in wb.sheets if sh.name.strip().upper() == 'FAM6_ADJ'), None)
        column_a = sim_sheet.range('A:A').value
        start_row = None
        last_row = None
        for i, value in enumerate(column_a, start=1):
            if value is not None and value != "":
                start_row = i
                break
        for i, value in enumerate(column_a[::-1], start=1):
            if value is not None and value != "":
                last_row = len(column_a) - i + 1
                break
        if start_row is None or last_row is None:
            raise RuntimeError("Simulation 시트에서 데이터 시작/종료 행을 찾지 못했습니다.")
        if last_row < start_row:
            raise RuntimeError(f"Simulation 범위 오류: start_row={start_row}, last_row={last_row}")
        df = sim_sheet.range(f'BB{start_row}:CJ{last_row}').options(pd.DataFrame, index=False, header=False).value
        drop_cols = [c for c in [33, 34] if c in df.columns]
        if drop_cols:
            df.drop(drop_cols, axis=1, inplace=True)
        if df.shape[1] != len(SIM_COLUMNS):
            raise RuntimeError(f'Simulation 컬럼 수 불일치: {df.shape[1]} / 기대 {len(SIM_COLUMNS)}')
        df.columns = SIM_COLUMNS
        if fam6_sheet is None:
            self.logger.warning('FAM6_ADJ 시트를 찾지 못했습니다. UI FAM6 매핑을 사용합니다.')
            df_fam6 = pd.DataFrame(columns=['FAM6', 'FAM6_ADJ'])
        else:
            df_fam6 = fam6_sheet.range('A1').options(pd.DataFrame, index=False, expand='table').value
            if 'FAM6' not in df_fam6.columns or 'FAM6_ADJ' not in df_fam6.columns:
                raise RuntimeError('FAM6_ADJ 시트에 FAM6 / FAM6_ADJ 컬럼이 필요합니다.')
        return df, df_fam6
    @staticmethod
    def close_workbook_safe(wb):
        try:
            wb.close()
        except Exception:
            pass

class DataTransformer:
    def __init__(self, logger):
        self.logger = logger

    @staticmethod
    def _normalize_line_value(v):
        if pd.isna(v):
            return v
        s = str(v).strip()
        if s.endswith('.0'):
            head = s[:-2]
            if head.isdigit():
                s = head
        return s

    def prepare_raw_df(self, df, df_fam6, planid):
        df = df[df['GUBUN'].isin(GUBUN_VALUES)].copy().reset_index(drop=True)
        df['LINE'] = df['LINE'].apply(self._normalize_line_value)
        for col in ['PRE_M9', 'PRE_M10', 'PRE_M11', 'PRE_M12']:
            df[col] = 0
        df['PLANID'] = planid
        df = df[['PLANID'] + [c for c in df.columns if c != 'PLANID']]
        df_fam6 = df_fam6.drop_duplicates(subset=['FAM6'], keep='first')
        df = pd.merge(df, df_fam6, on='FAM6', how='left')
        if 'FAM6_ADJ' not in df.columns:
            raise RuntimeError('FAM6 merge 후 FAM6_ADJ 컬럼이 없습니다.')
        return df
class ExcelProcessor:
    def __init__(self, logger):
        self.logger = logger
        self.transformer = DataTransformer(logger)

    def build_summary_files(self, df_sunipgo, df_info, output_dir):
        df_sunipgo = df_sunipgo.copy()
        df_sunipgo['LINE'] = df_sunipgo['LINE'].apply(DataTransformer._normalize_line_value)
        df_sunipgo2 = df_sunipgo.merge(DF_LINE, how='left', on='LINE')
        df_sunipgo2 = df_sunipgo2.merge(DF_DR, how='left', on='DESIGN_RULE')
        df_sunipgo2['LINE2'] = pd.Categorical(df_sunipgo2['LINE2'], categories=ORDER_LINE3, ordered=True)
        df_sunipgo2['DR2'] = pd.Categorical(df_sunipgo2['DR2'], categories=ORDER_DR2, ordered=True)
        df_sunipgo2 = df_sunipgo2.sort_values(by=['FAM1', 'LINE2', 'DR2'], ascending=[True, True, True])
        df_wf_hbm = pd.pivot_table(df_sunipgo2[df_sunipgo2['FAM6'].astype(str).str.contains('HBM', na=False)], values='WF_TTL', index=['FAM1', 'DR2'], columns='YEARMONTH', aggfunc='sum', fill_value=0, margins=True, observed=True)
        df_wf_hbm = df_wf_hbm[df_wf_hbm['All'] != 0].drop(index='All', errors='ignore').drop(columns='All', errors='ignore')
        columns_2026 = [col for col in df_wf_hbm.columns if str(col).startswith('2026')]
        columns_2027 = [col for col in df_wf_hbm.columns if str(col).startswith('2027')]
        df_wf_hbm['2026합'] = df_wf_hbm[columns_2026].sum(axis=1)
        df_wf_hbm['2027합'] = df_wf_hbm[columns_2027].sum(axis=1)
        df_wf_hbm = pd.concat([df_wf_hbm[columns_2026], df_wf_hbm[['2026합']], df_wf_hbm[columns_2027], df_wf_hbm[['2027합']]], axis=1).reset_index()
        df_wf_hbm['FAM1'] = 'HBM'
        df_wf_all = make_df_pv(df_sunipgo2, col_value='WF_TTL', order_line=ORDER_LINE3, colsum_true=True)
        df_wf_er = make_df_pv(df_sunipgo2[df_sunipgo2['FAM6'].astype(str).str.contains('ER', na=False)], col_value='WF_TTL', order_line=ORDER_LINE3, colsum_true=False)
        df_pv_sunipgo = make_df_pv(df_sunipgo2, col_value='순입고_억EQ', order_line=ORDER_LINE3, colsum_true=True)
        df_pv_sunsangsan = make_df_pv(df_sunipgo2, col_value='순생산_억EQ', order_line=ORDER_LINE3, colsum_true=True)
        df_sunipgo2 = df_sunipgo2.merge(DF_DR2, how='left', on='DESIGN_RULE')
        df_sunipgo2['DR3'] = pd.Categorical(df_sunipgo2['DR3'], categories=ORDER_DR2, ordered=True)
        df_sunipgo2 = df_sunipgo2.sort_values(by=['FAM1', 'LINE2', 'DR3'], ascending=[True, True, True])
        df_summary = pd.pivot_table(df_sunipgo2, values=['WF_TTL', '순입고_억EQ', '순생산_억EQ', 'PKG입고_억EQ'], index=['FAM1', 'DR3'], columns='YEAR', aggfunc='sum', fill_value=0, margins=False, observed=True)
        custom_order = [('WF_TTL', '2026'), ('순입고_억EQ', '2026'), ('순생산_억EQ', '2026'), ('PKG입고_억EQ', '2026'), ('WF_TTL', '2027'), ('순입고_억EQ', '2027'), ('순생산_억EQ', '2027'), ('PKG입고_억EQ', '2027')]
        custom_order = [c for c in custom_order if c in df_summary.columns]
        df_summary = df_summary[custom_order]
        wf_mask = df_summary.columns.get_level_values(0).str.contains('WF_TTL', case=False)
        df_summary.loc[:, wf_mask] = df_summary.loc[:, wf_mask] / 1000
        subtotals = df_summary.groupby(level=['FAM1']).sum()
        subtotals.index = pd.MultiIndex.from_tuples([(idx[0], '합계') for idx in subtotals.index])
        df_summary = pd.concat([subtotals, df_summary]).swaplevel(0, 1, axis=1)
        df_sunipgo2['MONTH'] = df_sunipgo2['YEARMONTH'].astype(str).str[-2:].astype(int)
        df_sunipgo2['QUARTER'] = df_sunipgo2['MONTH'].apply(month_to_quarter)
        df_pkg_dram = pd.pivot_table(df_sunipgo2[df_sunipgo2['FAM1'] == 'DRAM'], values=['PKG입고_억EQ'], index=['FAM1', 'DR3'], columns=['YEAR', 'QUARTER'], aggfunc='sum', fill_value=0, observed=True)
        subtotals = df_pkg_dram.groupby(level=['FAM1'], observed=True).sum()
        subtotals.index = pd.MultiIndex.from_tuples([(idx[0], '합계') for idx in subtotals.index])
        df_pkg_dram = pd.concat([subtotals, df_pkg_dram])
        if not df_pkg_dram.empty:
            df_pkg_dram['2026'] = df_pkg_dram.iloc[:, 0:4].sum(axis=1)
            df_pkg_dram['2027'] = df_pkg_dram.iloc[:, 4:8].sum(axis=1)
        df_pkg_dram_hbm = pd.pivot_table(df_sunipgo2[(df_sunipgo2['FAM1'] == 'DRAM') & (df_sunipgo2['FAM6'].astype(str).str.contains('HBM', na=False))], values=['PKG입고_억EQ'], index=['FAM1', 'DR3'], columns=['YEAR', 'QUARTER'], aggfunc='sum', fill_value=0, observed=True)
        subtotals = df_pkg_dram_hbm.groupby(level=['FAM1'], observed=True).sum()
        subtotals.index = pd.MultiIndex.from_tuples([(idx[0], '합계') for idx in subtotals.index])
        df_pkg_dram_hbm = pd.concat([subtotals, df_pkg_dram_hbm])
        if not df_pkg_dram_hbm.empty:
            df_pkg_dram_hbm['2026'] = df_pkg_dram_hbm.iloc[:, 0:4].sum(axis=1)
            df_pkg_dram_hbm['2027'] = df_pkg_dram_hbm.iloc[:, 4:8].sum(axis=1)
        df_pkg_dram = pd.concat([df_pkg_dram, df_pkg_dram_hbm])
        df_sunipgo2 = pd.merge(df_sunipgo2, df_info[['FAM6', 'USER_FAM1', 'USER_FAM2']], on='FAM6', how='left')
        df_pkg_flash = pd.pivot_table(df_sunipgo2[(df_sunipgo2['FAM1'] == 'FLASH') & (df_sunipgo2['VERSION'] != '-')], values=['PKG입고_억EQ'], index=['FAM1', 'DR3', 'VERSION'], columns=['YEAR', 'QUARTER'], aggfunc='sum', fill_value=0, observed=True)
        if not df_pkg_flash.empty:
            df_pkg_flash = df_pkg_flash.reset_index()
        df_psi = pd.pivot_table(df_sunipgo2[df_sunipgo2['YEAR'].astype(str).str.contains('2027', na=False)], values=['WF_TTL', '순생산_억EQ', '순입고_억EQ', 'PKG입고_억EQ'], index=['FAM1', 'DR2', 'VERSION'], columns='YEARMONTH', aggfunc='sum', fill_value=0, observed=True)
        measures = ['WF_TTL', '순생산_억EQ', '순입고_억EQ', 'PKG입고_억EQ']
        dates = df_psi.columns.get_level_values('YEARMONTH').unique().values
        custom_order = [(measure, date) for measure in measures for date in dates if (measure, date) in df_psi.columns]
        df_psi = df_psi[custom_order]
        wf_mask = df_psi.columns.get_level_values(0).str.contains('WF_TTL', case=False)
        df_psi.loc[:, wf_mask] = df_psi.loc[:, wf_mask] / 1000
        df_dr_psi = df_psi.loc[:, wf_mask].groupby(level=['FAM1', 'DR2'], observed=True).sum()
        summary_path = output_dir / 'summary.xlsx'
        with pd.ExcelWriter(summary_path, engine='xlsxwriter') as writer:
            df_wf_hbm.to_excel(writer, sheet_name='WF', index=False, startrow=0)
            df_wf_all.to_excel(writer, sheet_name='WF', index=False, startrow=len(df_wf_hbm) + 2)
            df_wf_er.to_excel(writer, sheet_name='WF', index=False, startrow=len(df_wf_all) + len(df_wf_hbm) + 4)
            df_pv_sunipgo.to_excel(writer, sheet_name='Sunipgo', index=False, startrow=0)
            df_pv_sunsangsan.to_excel(writer, sheet_name='Sunipgo', index=False, startrow=len(df_pv_sunipgo) + 2)
            df_summary.to_excel(writer, sheet_name='계수', startrow=0)
            df_pkg_dram.to_excel(writer, sheet_name='PKG입고', startrow=0)
            df_pkg_flash.to_excel(writer, sheet_name='PKG입고', startrow=0, startcol=12)
            df_psi.to_excel(writer, sheet_name='psi', startrow=1)
            df_dr_psi.to_excel(writer, sheet_name='psi', startrow=1, startcol=df_psi.shape[1] + 4)
        wb = load_workbook(summary_path)
        ws0 = wb['WF']
        ws0[f'A{len(df_wf_all) + len(df_wf_hbm) + 4}'] = 'ER'
        ws1 = wb['Sunipgo']
        ws1[f'A{len(df_pv_sunipgo) + 2}'] = '순생산'
        ws2 = wb['psi']
        for i in range(49):
            ws2[f'{get_column_letter(3 + i)}1'] = i + 1
        for i in range(13):
            ws2[f'{get_column_letter(54 + i)}1'] = i + 1
        wb.save(summary_path)
        return summary_path

    def _build_o9_body(self, df_o9: pd.DataFrame, planid: str) -> pd.DataFrame:
        value_cols = [f"Y0M{i}" for i in range(1, 13)] + [f"Y1M{i}" for i in range(1, 13)]
        missing = [c for c in ["VERSION_NAME", "FAM6", "SITEID", "MEASURE"] + value_cols if c not in df_o9.columns]
        if missing:
            raise RuntimeError(f"O9 결과 생성용 컬럼 누락: {missing}")
        grouped = (
            df_o9[["VERSION_NAME", "FAM6", "SITEID", "MEASURE"] + value_cols]
            .groupby(["VERSION_NAME", "FAM6", "SITEID", "MEASURE"], dropna=False, as_index=False)
            .sum(numeric_only=True)
        )
        grouped = grouped.rename(columns={"MEASURE": "Measure"})
        o9_months = build_o9_months(planid)
        rename_map = {col: month for col, month in zip(value_cols, o9_months)}
        grouped = grouped.rename(columns=rename_map)
        ordered_cols = ["VERSION_NAME", "FAM6", "SITEID", "Measure"] + o9_months
        grouped = grouped[ordered_cols]
        for col in o9_months:
            grouped[col] = pd.to_numeric(grouped[col], errors="coerce").fillna(0.0)
        return grouped

    def build_o9_upload_file(self, df_o9: pd.DataFrame, output_dir: Path, planid: str) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"o9_upload_{sanitize_filename(planid)}_{timestamp}.xlsx"
        body_df = self._build_o9_body(df_o9, planid)
        o9_months = body_df.columns.tolist()[4:]

        year_header = ["", "", "", ""] + [f"Time.[Year].[{m[:4]}]" for m in o9_months]
        month_header = ["", "", "", ""] + [f"Time.[Month].[{m}]" for m in o9_months]
        mtd_header = ["Version.[Version Name]", "FAM6Item.[FAM6]", "Line.[Line]", "Measure"] + [
            f"Time.[MTD].[{calendar.monthrange(int(m[:4]), int(m[4:]))[1]}]" for m in o9_months
        ]
        header_df = pd.DataFrame([year_header, month_header, mtd_header], columns=body_df.columns)
        final_df = pd.concat([header_df, body_df], ignore_index=True)

        with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
            final_df.to_excel(writer, sheet_name="o9_final", index=False, header=False)
            ws = writer.sheets["o9_final"]
            num_fmt = writer.book.add_format({"num_format": "0.0000"})
            start_col = 4  # E열부터 숫자영역
            end_col = start_col + len(o9_months) - 1
            ws.set_column(start_col, end_col, 14, num_fmt)
        self.logger.info("O9 업로드 파일 생성: %s", output_path)
        return output_path

class JobRunner:
    def __init__(self, app, logger):
        self.app = app
        self.logger = logger

    def run(self, edm_link, planid, keep_excel_open=False):
        start_time = datetime.now()
        wb = None
        oracle = OracleService(self.logger)
        try:
            output_dir = Path.cwd() / 'output' / f"{datetime.now():%Y%m%d_%H%M%S}_{sanitize_filename(planid)}"
            output_dir.mkdir(parents=True, exist_ok=True)
            self.app.last_output_dir = output_dir
            self.logger.info('Output 폴더: %s', output_dir)
            reader = EDMClient(self.logger)
            wb = reader.open_edm_and_attach_workbook(edm_link)
            df_raw, df_fam6 = reader.read_simulation_and_fam6(wb)
            ui_mapping = self.app.get_fam6_mapping_df()
            if not ui_mapping.empty:
                df_fam6 = ui_mapping
                self.logger.info("UI FAM6 매핑 사용: %s건", len(df_fam6))
            if not keep_excel_open:
                reader.close_workbook_safe(wb)
                wb = None
            transformer = DataTransformer(self.logger)
            df = transformer.prepare_raw_df(df_raw, df_fam6, planid)
            with pd.ExcelWriter(output_dir / 'ebaTOgscmdb.xlsx', engine='openpyxl') as writer:
                df.to_excel(writer, index=False, header=True, startrow=0, startcol=0)
            oracle.connect()
            oracle.delete_table_if_exists(CFG.NEW_TABLE_NAME)
            oracle.copy_table_structure(CFG.OLD_TABLE_NAME, CFG.NEW_TABLE_NAME)
            oracle.insert_dataframe_into_table(df, CFG.NEW_TABLE_NAME)
            params = {'plan_id': planid}
            df_sunipgo = oracle.read_sql(QUERY, params=params)
            df_info = oracle.read_sql(QUERY2)
            df_sunipgo.to_excel(output_dir / 'df_sunipgo.xlsx', index=False)
            excel_processor = ExcelProcessor(self.logger)
            excel_processor.build_summary_files(df_sunipgo, df_info, output_dir)
            df_o9 = oracle.read_sql(O9_BASE_QUERY, params=params)
            excel_processor.build_o9_upload_file(df_o9, output_dir, planid)
            self.logger.info(elapsed_text(start_time))
            self.app.after(0, lambda: messagebox.showinfo('완료', f'작업이 완료되었습니다.\n\nOutput 폴더:\n{output_dir}'))
        except Exception as e:
            self.logger.error("오류 요약:\n%s", traceback.format_exc(limit=5))
            self.logger.exception('실행 중 오류 발생')
            self.app.after(0, lambda: messagebox.showerror('오류', f'작업 중 오류가 발생했습니다.\n\n{e}'))
        finally:
            try:
                oracle.close()
            except Exception:
                pass
            if wb is not None and not keep_excel_open:
                try:
                    wb.close()
                except Exception:
                    pass
            self.app.after(0, self.app._finish_run)


class Fam6ManagerDialog(tk.Toplevel):
    def __init__(self, master, df_mapping: pd.DataFrame, on_save):
        super().__init__(master)
        self.title("FAM6 관리")
        self.geometry("760x520")
        self.on_save = on_save
        self.tree = ttk.Treeview(self, columns=("FAM6", "FAM6_ADJ"), show="headings")
        self.tree.heading("FAM6", text="FAM6")
        self.tree.heading("FAM6_ADJ", text="FAM6_ADJ")
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        for _, row in df_mapping.fillna("").iterrows():
            self.tree.insert("", "end", values=(row["FAM6"], row["FAM6_ADJ"]))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=10)
        self.fam6_var = tk.StringVar()
        self.fam6_adj_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.fam6_var, width=35).pack(side="left", padx=4)
        ttk.Entry(form, textvariable=self.fam6_adj_var, width=35).pack(side="left", padx=4)
        ttk.Button(form, text="추가", command=self.add_row).pack(side="left", padx=4)
        ttk.Button(form, text="선택삭제", command=self.delete_selected).pack(side="left", padx=4)
        ttk.Button(form, text="저장", command=self.save).pack(side="right", padx=4)

    def add_row(self):
        fam6 = self.fam6_var.get().strip()
        fam6_adj = self.fam6_adj_var.get().strip()
        if fam6:
            self.tree.insert("", "end", values=(fam6, fam6_adj or fam6))
            self.fam6_var.set("")
            self.fam6_adj_var.set("")

    def delete_selected(self):
        for item in self.tree.selection():
            self.tree.delete(item)

    def save(self):
        rows = [self.tree.item(i, "values") for i in self.tree.get_children("")]
        df = pd.DataFrame(rows, columns=["FAM6", "FAM6_ADJ"])
        df = df[(df["FAM6"].astype(str).str.strip() != "")]
        self.on_save(df)
        self.destroy()


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('EBA → GSCM DB / Summary Tool')
        self.geometry('980x700')
        self.log_queue = queue.Queue()
        self.worker = None
        self.last_output_dir = None
        self.settings = load_settings()
        self.fam6_mapping_df = load_fam6_mapping_file()
        self.logger = logging.getLogger('eba_tool')
        self.logger.setLevel(logging.INFO)
        self.logger.handlers.clear()
        qh = QueueLogHandler(self.log_queue)
        qh.setFormatter(logging.Formatter('%(asctime)s | %(levelname)s | %(message)s', '%H:%M:%S'))
        self.logger.addHandler(qh)
        self._build_ui()
        self.after(200, self._drain_log_queue)
    def _build_ui(self):
        top = ttk.Frame(self, padding=12)
        top.pack(fill='x')
        ttk.Label(top, text='EDM LINK').grid(row=0, column=0, sticky='w', padx=(0, 8), pady=4)
        self.edm_var = tk.StringVar(value=self.settings.get("edm_link", CFG.DEFAULT_EDM_LINK))
        ttk.Entry(top, textvariable=self.edm_var, width=110).grid(row=0, column=1, sticky='ew', pady=4)
        ttk.Label(top, text='PLANID').grid(row=1, column=0, sticky='w', padx=(0, 8), pady=4)
        self.planid_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.planid_var, width=40).grid(row=1, column=1, sticky='w', pady=4)
        self.keep_excel_open_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(top, text='작업 후 Excel 워크북 닫지 않기', variable=self.keep_excel_open_var).grid(row=2, column=1, sticky='w', pady=(4, 8))
        btns = ttk.Frame(top)
        btns.grid(row=3, column=1, sticky='w')
        self.run_btn = ttk.Button(btns, text='실행', command=self.start_run)
        self.run_btn.pack(side='left')
        ttk.Button(btns, text='로그 지우기', command=lambda: self.log_text.delete('1.0', 'end')).pack(side='left', padx=6)
        ttk.Button(btns, text='Output 폴더 열기', command=self.open_output_folder).pack(side='left')
        ttk.Button(btns, text='FAM6 관리', command=self.open_fam6_manager).pack(side='left', padx=6)
        top.columnconfigure(1, weight=1)
        mid = ttk.Frame(self, padding=(12, 0, 12, 8))
        mid.pack(fill='x')
        self.status_var = tk.StringVar(value='대기 중')
        ttk.Label(mid, textvariable=self.status_var).pack(anchor='w')
        body = ttk.Frame(self, padding=(12, 0, 12, 12))
        body.pack(fill='both', expand=True)
        self.log_text = tk.Text(body, wrap='word', font=('Consolas', 10))
        self.log_text.pack(side='left', fill='both', expand=True)
        scroll = ttk.Scrollbar(body, orient='vertical', command=self.log_text.yview)
        scroll.pack(side='right', fill='y')
        self.log_text.config(yscrollcommand=scroll.set)
    def open_output_folder(self):
        if not self.last_output_dir:
            messagebox.showinfo('안내', '아직 생성된 Output 폴더가 없습니다.')
            return
        os.startfile(str(self.last_output_dir))
    def start_run(self):
        if self.worker and self.worker.is_alive():
            messagebox.showwarning('실행 중', '이미 작업이 실행 중입니다.')
            return
        planid = self.planid_var.get().strip()
        edm_link = self.edm_var.get().strip()
        if not planid:
            messagebox.showwarning('입력 필요', 'PLANID를 입력해 주세요.')
            return
        if not edm_link:
            messagebox.showwarning('입력 필요', 'EDM LINK를 입력해 주세요.')
            return
        self.settings["edm_link"] = edm_link
        save_settings(self.settings)
        self.run_btn.config(state='disabled')
        self.status_var.set('실행 중...')
        runner = JobRunner(self, self.logger)
        self.worker = threading.Thread(target=runner.run, args=(edm_link, planid, self.keep_excel_open_var.get()), daemon=True)
        self.worker.start()

    def open_fam6_manager(self):
        Fam6ManagerDialog(self, self.fam6_mapping_df, self._save_fam6_mapping)

    def _save_fam6_mapping(self, df: pd.DataFrame):
        self.fam6_mapping_df = df.drop_duplicates(subset=["FAM6"], keep="first")
        save_fam6_mapping_file(self.fam6_mapping_df)
        self.logger.info("FAM6 매핑 저장 완료: %s건", len(self.fam6_mapping_df))

    def get_fam6_mapping_df(self) -> pd.DataFrame:
        return self.fam6_mapping_df.copy()
    def _finish_run(self):
        self.run_btn.config(state='normal')
        self.status_var.set('대기 중')
    def _drain_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                self.log_text.insert('end', msg + '\n')
                self.log_text.see('end')
        except queue.Empty:
            pass
        finally:
            self.after(200, self._drain_log_queue)

def main():
    app = App()
    app.mainloop()

if __name__ == '__main__':
    main()
